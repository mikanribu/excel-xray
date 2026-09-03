"""Generate a deliberately complicated EUC workbook for manual testing.

Run with openpyxl available (it is a build-time tool only, not a runtime dep):

    uv run --with openpyxl python tests/fixtures/make_complex_fixture.py

Produces `complex_euc_model.xlsx` next to this script. It exercises the whole
pipeline: several tab types (input / mapping / calculation / control check /
reporting / output / supporting), a ListObject, a merged multi-row header,
named ranges, cross-sheet dependencies, lookups + reconciliation with a
tolerance, volatile functions, hardcoded inputs, data validation, and a hidden
leftover sheet — plus real error-producing formulas (=1/0, VLOOKUP of a missing
key).

Note: openpyxl does not compute formulas, so the file has no *cached* values.
The tool will say so, and value-derived signals (cached errors, number/date
typing of formula cells) stay quiet until the file is opened and saved once in
Excel, which bakes the cached results (including the #DIV/0! and #N/A above).
Everything structural — formulas, skeletons, categories, dependencies, volatile
functions, tables, merges, named ranges, the hidden sheet — works immediately.
"""

from __future__ import annotations

import datetime as dt
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.table import Table, TableStyleInfo

HDR = Font(bold=True, color="FFFFFF")
HDR_FILL = PatternFill("solid", fgColor="2D6CA2")
BOLD = Font(bold=True)


def _header(ws, row, values, start_col=1):
    for i, v in enumerate(values):
        c = ws.cell(row=row, column=start_col + i, value=v)
        c.font = HDR
        c.fill = HDR_FILL


def build() -> Workbook:
    wb = Workbook()

    # ---- 1. Instructions (supporting / notes) --------------------------
    ws = wb.active
    ws.title = "Instructions"
    ws["A1"] = "Quarterly Reserving & MI Model (TEST FIXTURE)"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A3"] = (
        "Purpose: consolidate policy-level premiums, apply reserve factors by "
        "line of business, reconcile to source, and produce the quarterly MI "
        "pack. Owner: Reserving team. Prepared month-end + 3 working days. This "
        "cell is a long free-text note so the detector classifies this block as "
        "notes rather than a table."
    )

    # ---- 2. Assumptions (input / key-value) ----------------------------
    a = wb.create_sheet("Assumptions")
    a["A1"], a["B1"] = "Parameter", "Value"
    a["A1"].font = a["B1"].font = BOLD
    params = [
        ("Discount rate", 0.04),
        ("Tax rate", 0.19),
        ("Reserve tolerance (GBP)", 1000),
        ("FX USD/GBP", 0.79),
        ("FX EUR/GBP", 0.86),
        ("Loss ratio cap", 0.85),
        ("Reporting date", "=TODAY()"),          # volatile
    ]
    for i, (k, v) in enumerate(params, start=2):
        a.cell(row=i, column=1, value=k)
        a.cell(row=i, column=2, value=v)
    wb.defined_names["DiscountRate"] = DefinedName("DiscountRate", attr_text="Assumptions!$B$2")
    wb.defined_names["TaxRate"] = DefinedName("TaxRate", attr_text="Assumptions!$B$3")
    wb.defined_names["Tolerance"] = DefinedName("Tolerance", attr_text="Assumptions!$B$4")

    # ---- 3. Raw_Premiums (input, ListObject) ---------------------------
    r = wb.create_sheet("Raw_Premiums")
    cols = ["Policy ID", "LOB Code", "Region", "Currency",
            "Written Premium", "Earned Premium", "Inception Date"]
    _header(r, 1, cols)
    lobs = ["MOT", "PRO", "LIA", "MAR"]
    regions = ["North", "South", "East", "West"]
    ccy = ["GBP", "USD", "EUR"]
    base = dt.date(2021, 1, 1)
    n_rows = 2500                                    # realistic volume -> Medium
    for i in range(n_rows):
        row = i + 2
        r.cell(row, 1, f"POL{1000 + i}")
        r.cell(row, 2, lobs[i % len(lobs)])
        r.cell(row, 3, regions[i % len(regions)])
        r.cell(row, 4, ccy[i % len(ccy)])
        written = 100000 + (i * 7919) % 900000
        r.cell(row, 5, written)
        r.cell(row, 6, f"=E{row}*0.85")           # earned = 85% of written (hardcoded)
        r.cell(row, 7, base + dt.timedelta(days=i * 11))
    last = n_rows + 1
    tbl = Table(displayName="tblPremiums", ref=f"A1:G{last}")
    tbl.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2", showRowStripes=True)
    r.add_table(tbl)
    dv = DataValidation(type="list", formula1='"North,South,East,West"', allow_blank=True)
    r.add_data_validation(dv)
    dv.add(f"C2:C{last}")

    # ---- 4. LOB_Mapping (mapping / lookup) -----------------------------
    m = wb.create_sheet("LOB_Mapping")
    _header(m, 1, ["LOB Code", "LOB Name", "Reserve Factor"])
    for i, (code, name, f) in enumerate([
        ("MOT", "Motor", 0.65), ("PRO", "Property", 0.55),
        ("LIA", "Liability", 0.72), ("MAR", "Marine", 0.60),
    ], start=2):
        m.cell(i, 1, code)
        m.cell(i, 2, name)
        m.cell(i, 3, f)

    # ---- 5. FX_Rates (mapping) -----------------------------------------
    fx = wb.create_sheet("FX_Rates")
    _header(fx, 1, ["Currency", "Rate to GBP"])
    for i, (c, v) in enumerate([("GBP", 1.0), ("USD", "=DiscountRate*0+0.79"),
                                ("EUR", 0.86)], start=2):
        fx.cell(i, 1, c)
        fx.cell(i, 2, v)

    # ---- 6. Calc_Reserves (calculation, merged multi-row header) -------
    c = wb.create_sheet("Calc_Reserves")
    c["A1"] = "Reserve Roll-Forward"
    c["A1"].font = Font(bold=True, size=12)
    # merged parent band on row 3, children on row 4
    c["B3"], c["D3"], c["F3"] = "Gross", "Reinsurance", "Net"
    for cell in ("B3", "D3", "F3"):
        c[cell].font = HDR
        c[cell].fill = HDR_FILL
    c.merge_cells("B3:C3")
    c.merge_cells("D3:E3")
    c.merge_cells("F3:G3")
    _header(c, 4, ["Accident Year", "Opening", "Movement",
                   "Opening", "Movement", "Opening", "Movement"])
    for i, year in enumerate(range(2019, 2025)):
        row = 5 + i
        c.cell(row, 1, year)
        c.cell(row, 2, f'=SUMIF(Raw_Premiums!$C:$C,"North",Raw_Premiums!$F:$F)*(1+DiscountRate)')
        c.cell(row, 3, f"=B{row}*0.3")                                   # hardcoded
        c.cell(row, 4, f'=B{row}*VLOOKUP("MOT",LOB_Mapping!$A:$C,3,FALSE)')  # lookup
        c.cell(row, 5, f"=D{row}*0.3")
        c.cell(row, 6, f"=B{row}-D{row}")
        c.cell(row, 7, f"=C{row}-E{row}")
    tr = 11
    c.cell(tr, 1, "Total").font = BOLD
    for col in range(2, 8):
        letter = chr(64 + col)
        cell = c.cell(tr, col, f"=SUM({letter}5:{letter}10)")
        cell.font = BOLD

    # ---- 7. Reconciliation (control check) -----------------------------
    rc = wb.create_sheet("Reconciliation")
    rc["A1"] = "Reconciliation check"
    rc["A1"].font = Font(bold=True, size=12)
    _header(rc, 3, ["Metric", "Model", "Source", "Variance", "Status"])
    rc["A4"] = "Total Earned (North)"
    rc["B4"] = "=Calc_Reserves!B11"
    rc["C4"] = '=SUMIF(Raw_Premiums!$C:$C,"North",Raw_Premiums!$F:$F)'
    rc["D4"] = "=B4-C4"
    rc["E4"] = '=IF(ABS(D4)<Tolerance,"OK","CHECK")'            # tolerance test
    rc["A5"] = "Missing LOB lookup"
    rc["B5"] = '=VLOOKUP("XXX",LOB_Mapping!$A:$C,2,FALSE)'      # -> #N/A when recalced
    rc["A6"] = "Coverage ratio"
    rc["B6"] = "=Calc_Reserves!F11/C6"                          # C6 blank -> #DIV/0!
    rc["A7"] = "Check date"
    rc["B7"] = "=NOW()"                                         # volatile

    # ---- 8. MI_Pack (reporting / output) -------------------------------
    mi = wb.create_sheet("MI_Pack")
    mi["A1"] = "Quarterly MI Pack"
    mi["A1"].font = Font(bold=True, size=12)
    _header(mi, 3, ["Region", "Q1", "Q2", "Q3", "Q4", "FY"])
    for i, reg in enumerate(regions, start=4):
        mi.cell(i, 1, reg)
        for q in range(4):
            mi.cell(i, 2 + q,
                    f'=SUMIF(Raw_Premiums!$C:$C,"{reg}",Raw_Premiums!$E:$E)/4')
        mi.cell(i, 6, f"=SUM(B{i}:E{i})")
    mi.cell(8, 1, "Total").font = BOLD
    for col in range(2, 7):
        letter = chr(64 + col)
        mi.cell(8, col, f"=SUM({letter}4:{letter}7)").font = BOLD

    # ---- 9. Dashboard (output, volatile OFFSET/INDIRECT) ---------------
    d = wb.create_sheet("Dashboard")
    d["A1"] = "Executive dashboard"
    d["A1"].font = Font(bold=True, size=12)
    d["A3"] = "Headline FY premium"
    d["B3"] = "=MI_Pack!F8"
    d["A4"] = "North FY (via OFFSET)"
    d["B4"] = "=OFFSET(MI_Pack!$A$4,0,5)"                       # volatile
    d["A5"] = "South FY (via INDIRECT)"
    d["B5"] = '=INDIRECT("MI_Pack!F5")'                         # volatile
    d["A6"] = "As at"
    d["B6"] = "=Assumptions!B8"

    # ---- 10. Old_Working_v1 (hidden leftover) --------------------------
    o = wb.create_sheet("Old_Working_v1")
    o.sheet_state = "hidden"
    o["A1"] = "superseded working - do not use"
    o["A3"] = "scratch"
    o["B3"] = "=Calc_Reserves!B5*1.1"
    o["B4"] = "=B3/0"                                           # -> #DIV/0!
    o["B5"] = "=B3+B4"

    return wb


if __name__ == "__main__":
    out = Path(__file__).with_name("complex_euc_model.xlsx")
    build().save(out)
    print(f"wrote {out}")
