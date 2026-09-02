"""Build a deliberately messy financial workbook as a test fixture.

Reproduces the structures that break naive extractors:
  - multi-row merged headers
  - two tables side by side on one sheet
  - a blank spacer row *inside* one logical table
  - a totals row that looks like data
  - a key-value assumptions block (not a table at all)
  - free prose notes in column A
  - a hidden working sheet
  - a real ListObject alongside hand-built ranges
  - a deliberate #DIV/0! left cached in the file
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import get_column_letter
import random

random.seed(7)

BOLD = Font(name="Arial", bold=True)
BLUE = Font(name="Arial", color="0000FF")
GREEN = Font(name="Arial", color="008000")
HDRFILL = PatternFill("solid", fgColor="DDDDDD")
BOT = Border(bottom=Side(style="thin"))

wb = Workbook()

# ---------------------------------------------------------------- Assumptions
ws = wb.active
ws.title = "Assumptions"
ws["A1"] = "Reserving Model - Key Assumptions"
ws["A1"].font = Font(name="Arial", bold=True, size=14)
ws["A3"] = "Valuation date"
ws["B3"] = "2025-06-30"
ws["B3"].font = BLUE
ws["A4"] = "Discount rate"
ws["B4"] = 0.035
ws["B4"].font = BLUE
ws["B4"].number_format = "0.0%"
ws["A5"] = "Expense loading"
ws["B5"] = 0.12
ws["B5"].font = BLUE
ws["B5"].number_format = "0.0%"
ws["A6"] = "Claims inflation"
ws["B6"] = 0.045
ws["B6"].font = BLUE
ws["B6"].number_format = "0.0%"
ws["A7"] = "FX rate HKD/USD"
ws["B7"] = 7.81
ws["B7"].font = BLUE

ws["A10"] = "Note: rates agreed with actuarial 14 Jul. Do not overwrite without sign-off."

# ---------------------------------------------------------------------- Data
ws = wb.create_sheet("Data")
# Left block: a real ListObject
hdr = ["Policy ID", "Line of Business", "Written Premium", "Earned Premium", "Region"]
for j, h in enumerate(hdr, start=1):
    c = ws.cell(row=1, column=j, value=h)
    c.font = BOLD
    c.fill = HDRFILL
lobs = ["Motor", "Property", "Marine", "Liability"]
regions = ["HK", "SG", "JP"]
for i in range(2, 42):
    ws.cell(row=i, column=1, value=f"POL-{10000+i}")
    ws.cell(row=i, column=2, value=random.choice(lobs))
    wp = round(random.uniform(5000, 90000), 2)
    ws.cell(row=i, column=3, value=wp)
    ws.cell(row=i, column=4, value=f"=C{i}*0.85")
    ws.cell(row=i, column=5, value=random.choice(regions))
tbl = Table(displayName="tblPolicies", ref="A1:E41")
tbl.tableStyleInfo = TableStyleInfo(name="TableStyleMedium9", showRowStripes=True)
ws.add_table(tbl)

# Right block: hand-built lookup table, side by side, separated by one blank col
ws["G1"] = "LOB Mapping"
ws["G1"].font = BOLD
for j, h in enumerate(["LOB Code", "LOB Name", "Reserve Factor"], start=7):
    c = ws.cell(row=2, column=j, value=h)
    c.font = BOLD
    c.border = BOT
for i, (code, name, f) in enumerate(
    [("MTR", "Motor", 0.62), ("PRP", "Property", 0.48),
     ("MAR", "Marine", 0.71), ("LIA", "Liability", 0.83)], start=3):
    ws.cell(row=i, column=7, value=code)
    ws.cell(row=i, column=8, value=name)
    ws.cell(row=i, column=9, value=f)

# ---------------------------------------------------------------------- Calc
ws = wb.create_sheet("Calc")
ws["A1"] = "Reserve Roll-Forward"
ws["A1"].font = Font(name="Arial", bold=True, size=12)

# multi-row merged header: parent spans two child columns each
ws.merge_cells("B3:C3")
ws["B3"] = "Gross"
ws.merge_cells("D3:E3")
ws["D3"] = "Reinsurance"
ws.merge_cells("F3:G3")
ws["F3"] = "Net"
for col in "BDF":
    ws[f"{col}3"].font = BOLD
    ws[f"{col}3"].fill = HDRFILL
ws["A4"] = "Accident Year"
for j, h in enumerate(["Opening", "Movement", "Opening", "Movement",
                       "Opening", "Movement"], start=2):
    c = ws.cell(row=4, column=j, value=h)
    c.font = BOLD
    c.border = BOT
ws["A4"].font = BOLD
ws["A4"].border = BOT

# block one: 2019-2022
for i, yr in enumerate(range(2019, 2023), start=5):
    ws.cell(row=i, column=1, value=str(yr))
    ws.cell(row=i, column=2, value=round(random.uniform(1e6, 5e6), 0))
    ws.cell(row=i, column=3, value=f"=B{i}*Assumptions!$B$6")
    ws.cell(row=i, column=4, value=f"=B{i}*0.3")
    ws.cell(row=i, column=5, value=f"=D{i}*Assumptions!$B$6")
    ws.cell(row=i, column=6, value=f"=B{i}-D{i}")
    ws.cell(row=i, column=7, value=f"=C{i}-E{i}")

# SPACER ROW at 9 -- same logical table continues
for i, yr in enumerate(range(2023, 2026), start=10):
    ws.cell(row=i, column=1, value=str(yr))
    ws.cell(row=i, column=2, value=round(random.uniform(1e6, 5e6), 0))
    ws.cell(row=i, column=3, value=f"=B{i}*Assumptions!$B$6")
    ws.cell(row=i, column=4, value=f"=B{i}*0.3")
    ws.cell(row=i, column=5, value=f"=D{i}*Assumptions!$B$6")
    ws.cell(row=i, column=6, value=f"=B{i}-D{i}")
    ws.cell(row=i, column=7, value=f"=C{i}-E{i}")

# totals row that looks like data
ws["A13"] = "Total"
ws["A13"].font = BOLD
for col in range(2, 8):
    L = get_column_letter(col)
    c = ws.cell(row=13, column=col,
                value=f"=SUM({L}5:{L}8)+SUM({L}10:{L}12)")
    c.font = BOLD

# a hardcoded constant buried in a calc block -- the thing we want to flag
ws["I5"] = "Adj factor"
ws["I6"] = 1.075
ws["I7"] = "=I6*1.02"

# deliberate divide-by-zero
ws["I9"] = 0
ws["I10"] = "=I6/I9"

ws["A16"] = "Notes:"
ws["A17"] = "2020 movement restated following audit query, see file Recon_2020_FINAL.xlsx"

# ------------------------------------------------------------------ MI Pack
ws = wb.create_sheet("MI Pack")
ws["A1"] = "Quarterly MI - Underwriting Result by Region"
ws["A1"].font = Font(name="Arial", bold=True, size=12)
ws["A3"] = "Region"
ws["A3"].font = BOLD
for j, q in enumerate(["Q1", "Q2", "Q3", "Q4", "FY"], start=2):
    c = ws.cell(row=3, column=j, value=q)
    c.font = BOLD
    c.border = BOT
for i, r in enumerate(["HK", "SG", "JP", "AU"], start=4):
    ws.cell(row=i, column=1, value=r)
    for j in range(2, 6):
        ws.cell(row=i, column=j, value=round(random.uniform(-5e5, 2e6), 0))
    ws.cell(row=i, column=6, value=f"=SUM(B{i}:E{i})")

# --------------------------------------------------------------- Old_Working
ws = wb.create_sheet("Old_Working")
ws["A1"] = "scratch - do not use"
for i in range(2, 20):
    ws.cell(row=i, column=1, value=random.random())
    ws.cell(row=i, column=2, value=f"=A{i}*2")
ws.sheet_state = "hidden"

wb.save("/home/claude/xray_project/samples/messy_reserving_model.xlsx")
print("written")
