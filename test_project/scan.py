"""Workbook X-ray: orchestrates structure, cells, regions and formulas.

Two openpyxl passes are unavoidable: data_only=False yields formula strings with
no results, data_only=True yields cached results with the formulas gone. One pass
cannot give you both, and both matter -- the formula is the logic, the cached
value is the evidence of what it last produced (including #REF! and #DIV/0!).

Known trap, detected and reported rather than swallowed: if the workbook was last
written by anything other than Excel, there are no cached values at all and
data_only=True returns None throughout.
"""
from __future__ import annotations

import datetime as _dt
import hashlib
import json
import os
from dataclasses import asdict, dataclass, field

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

from .formulas import ERROR_VALUES, FormulaProfile, analyse
from .ooxml import WorkbookStructure, read_structure
from .regions import CellFact, Region, detect_regions


@dataclass
class SheetXray:
    name: str
    position: int
    state: str
    dimension: str | None
    max_row: int
    max_col: int
    populated_cells: int
    density: float
    regions: list[Region] = field(default_factory=list)
    formula_profile: dict = field(default_factory=dict)
    merges: list[str] = field(default_factory=list)
    error_cells: list[str] = field(default_factory=list)
    notes: list[str] = field(default_factory=list)
    # Compact occupancy for the report plate: [row, col, flag] where flag is
    # 1 value, 2 formula, 3 error, 4 text. Structure only -- no cell values, so
    # the report can be shared without carrying confidential figures.
    occupancy: list[list[int]] = field(default_factory=list)
    occupancy_truncated: bool = False


@dataclass
class WorkbookXray:
    path: str
    filename: str
    size_bytes: int
    sha256: str
    fs_modified: str
    parse_status: str
    core_props: dict = field(default_factory=dict)
    app_props: dict = field(default_factory=dict)
    has_vba: bool = False
    has_power_query: bool = False
    external_links: list[str] = field(default_factory=list)
    connections: list[dict] = field(default_factory=list)
    pivot_cache_sources: list[str] = field(default_factory=list)
    defined_names: list = field(default_factory=list)
    sheets: list[SheetXray] = field(default_factory=list)
    warnings: list[str] = field(default_factory=list)


def _sha256(path: str, chunk: int = 1 << 20) -> str:
    h = hashlib.sha256()
    with open(path, "rb") as fh:
        for block in iter(lambda: fh.read(chunk), b""):
            h.update(block)
    return h.hexdigest()


def _cell_facts(fcell, vcell) -> CellFact | None:
    """Merge the formula-pass cell and value-pass cell into one fact."""
    raw = fcell.value
    if raw is None and (vcell is None or vcell.value is None):
        return None

    formula = None
    if isinstance(raw, str) and raw.startswith("="):
        formula = raw

    shown = vcell.value if (vcell is not None and vcell.value is not None) else (
        None if formula else raw)

    fact = CellFact(value=shown, formula=formula)
    if isinstance(shown, str):
        fact.is_text = not shown.startswith("=")
        fact.is_error = shown in ERROR_VALUES
        fact.text = shown
    elif isinstance(shown, (int, float)):
        fact.is_number = True
        fact.text = ""
    elif isinstance(shown, (_dt.datetime, _dt.date)):
        fact.is_date = True
        fact.text = ""
    elif shown is None and formula:
        fact.text = ""

    if fact.text == "" and isinstance(raw, str) and not formula:
        fact.is_text = True
        fact.text = raw

    try:
        fact.bold = bool(fcell.font and fcell.font.b)
        fill = fcell.fill
        fact.filled = bool(
            fill and fill.patternType and fill.patternType != "none")
        b = fcell.border
        fact.bordered = bool(
            b and ((b.bottom and b.bottom.style) or (b.top and b.top.style)))
    except Exception:
        pass
    return fact


OLE_MAGIC = b"\xd0\xcf\x11\xe0\xa1\xb1\x1a\xe1"
ZIP_MAGIC = b"PK\x03\x04"


class UnreadableWorkbook(Exception):
    """Raised with a human reason, so the coverage report is actionable."""

    def __init__(self, reason: str, category: str):
        super().__init__(reason)
        self.category = category


def triage(path: str) -> None:
    """Classify a file before parsing. Raises UnreadableWorkbook with a reason.

    'BadZipFile' tells an analyst nothing. 'Password-protected' tells them to go
    and ask for the password, which is the whole point of a coverage report.
    """
    size = os.path.getsize(path)
    if size == 0:
        raise UnreadableWorkbook("zero-byte file", "empty")
    with open(path, "rb") as fh:
        head = fh.read(8)

    if head.startswith(OLE_MAGIC):
        # OLE container: either a legacy .xls or an encrypted OOXML package.
        try:
            import olefile  # optional
            if olefile.isOleFile(path):
                ole = olefile.OleFileIO(path)
                streams = {"/".join(s) for s in ole.listdir()}
                ole.close()
                if "EncryptedPackage" in streams:
                    raise UnreadableWorkbook(
                        "password-protected (encrypted OOXML package)", "encrypted")
        except ImportError:
            pass
        raise UnreadableWorkbook(
            "legacy OLE format (.xls or encrypted) -- needs conversion first",
            "legacy_ole")

    if not head.startswith(ZIP_MAGIC):
        raise UnreadableWorkbook(
            f"not a zip container (magic {head[:4]!r}) -- extension may be wrong",
            "not_ooxml")

    import zipfile
    try:
        with zipfile.ZipFile(path) as zf:
            if "xl/workbook.xml" not in zf.namelist():
                raise UnreadableWorkbook(
                    "zip without xl/workbook.xml -- not a spreadsheet package",
                    "not_workbook")
    except zipfile.BadZipFile as e:
        raise UnreadableWorkbook(f"corrupt or truncated zip: {e}", "corrupt") from e


def xray_workbook(path: str, max_rows: int = 200_000) -> WorkbookXray:
    triage(path)
    st = os.stat(path)
    wx = WorkbookXray(
        path=os.path.abspath(path),
        filename=os.path.basename(path),
        size_bytes=st.st_size,
        sha256=_sha256(path),
        fs_modified=_dt.datetime.fromtimestamp(st.st_mtime).isoformat(timespec="seconds"),
        parse_status="full",
    )

    struct: WorkbookStructure = read_structure(path)
    wx.core_props = struct.core_props
    wx.app_props = struct.app_props
    wx.has_vba = struct.has_vba
    wx.has_power_query = struct.has_power_query
    wx.external_links = struct.external_links
    wx.connections = struct.connections
    wx.pivot_cache_sources = struct.pivot_cache_sources
    wx.defined_names = [{"name": n, "refers_to": r} for n, r in struct.defined_names]
    wx.warnings.extend(struct.parse_notes)

    app = (struct.app_props.get("application") or "")
    if app and "Excel" not in app:
        wx.warnings.append(
            f"last written by {app!r}, not Excel -- cached formula values may be "
            "absent, so value-derived signals are unreliable"
        )

    wb_f = load_workbook(path, read_only=True, data_only=False)
    wb_v = load_workbook(path, read_only=True, data_only=True)
    try:
        for sref in struct.sheets:
            if sref.name not in wb_f.sheetnames:
                wx.warnings.append(f"sheet {sref.name!r} not readable by openpyxl")
                wx.parse_status = "partial"
                continue
            sx = _xray_sheet(wb_f[sref.name], wb_v[sref.name], sref, struct, max_rows)
            wx.sheets.append(sx)
    finally:
        wb_f.close()
        wb_v.close()

    if all(s.populated_cells == 0 for s in wx.sheets) and wx.sheets:
        wx.parse_status = "partial"
        wx.warnings.append("no populated cells found in any sheet")
    return wx


def _xray_sheet(ws_f, ws_v, sref, struct, max_rows: int) -> SheetXray:
    sstruct = struct.structures.get(sref.name)
    merges = sstruct.merges if sstruct else []
    tables = sstruct.tables if sstruct else []

    cells: dict[tuple[int, int], CellFact] = {}
    profile = FormulaProfile()
    errors: list[str] = []
    notes: list[str] = []

    rows_f = ws_f.iter_rows()
    rows_v = ws_v.iter_rows()
    max_r = max_c = 0
    truncated = False

    # In read-only mode a row is a positional tuple padded with EmptyCell, which
    # carries no .row/.column. Coordinates therefore come from the iteration
    # index, not from the cell object.
    for i, row_f in enumerate(rows_f, start=1):
        if i > max_rows:
            truncated = True
            break
        row_v = next(rows_v, ())
        r = getattr(row_f[0], "row", i) if row_f else i
        for j, cf in enumerate(row_f, start=1):
            cv = row_v[j - 1] if j - 1 < len(row_v) else None
            fval = getattr(cf, "value", None)
            vval = getattr(cv, "value", None)
            if fval is None and vval is None:
                continue
            fact = _cell_facts(cf, cv)
            if fact is None:
                continue
            c = getattr(cf, "column", j) or j
            cells[(r, c)] = fact
            max_r, max_c = max(max_r, r), max(max_c, c)
            if fact.formula:
                profile.add(analyse(fact.formula, r, c))
            if fact.is_error:
                errors.append(f"{ws_f.title}!{get_column_letter(c)}{r} {fact.text}")

    if truncated:
        notes.append(f"row scan capped at {max_rows}")

    regions = detect_regions(cells, merges, tables)

    # Plate render caps: beyond this a miniature is unreadable anyway.
    PLATE_ROWS, PLATE_COLS = 120, 60
    occ: list[list[int]] = []
    occ_trunc = False
    for (r, c), f in cells.items():
        if r > PLATE_ROWS or c > PLATE_COLS:
            occ_trunc = True
            continue
        flag = 3 if f.is_error else 2 if f.formula else 4 if f.is_text else 1
        occ.append([r, c, flag])
    occ.sort()

    area = max_r * max_c
    return SheetXray(
        name=sref.name,
        position=sref.position,
        state=sref.state,
        dimension=sstruct.dimension if sstruct else None,
        max_row=max_r,
        max_col=max_c,
        populated_cells=len(cells),
        density=round(len(cells) / area, 4) if area else 0.0,
        regions=regions,
        merges=merges,
        error_cells=errors,
        notes=notes,
        occupancy=occ,
        occupancy_truncated=occ_trunc,
        formula_profile={
            "total": profile.total,
            "distinct_skeletons": profile.distinct_skeletons,
            "compression": round(profile.compression, 1),
            "volatile_count": profile.volatile_count,
            "cross_sheet_count": profile.cross_sheet_count,
            "external_count": profile.external_count,
            "hardcoded_literal_count": profile.hardcoded_literal_count,
            "referenced_sheets": dict(profile.referenced_sheets),
            "top_functions": profile.functions.most_common(12),
            "top_skeletons": profile.top(15),
        },
    )


def to_json(wx: WorkbookXray) -> str:
    def enc(o):
        if isinstance(o, (_dt.datetime, _dt.date)):
            return o.isoformat()
        if isinstance(o, set):
            return sorted(o)
        return str(o)
    d = asdict(wx)
    for s in d["sheets"]:
        for r in s["regions"]:
            r["ref"] = f"{r['top']},{r['left']}-{r['bottom']},{r['right']}"
    return json.dumps(d, indent=2, default=enc)


if __name__ == "__main__":
    import sys
    wx = xray_workbook(sys.argv[1])
    print(to_json(wx))
